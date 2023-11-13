import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime
class TelegramBot:
    def __init__(self, token):
        self.token = token
        self.url = f"https://api.telegram.org/bot{token}/"
        
    def get_updates(self):
        response = requests.get(f"{self.url}getUpdates")
        data = response.json()
        return data
    
    def get_last(self):
        data = self.get_updates()
        return data["result"][-1]["message"]["text"]
    
    def send_message(self, chat_id, text):
        requests.get(f"{self.url}sendMessage?chat_id={chat_id}&text={text}")

    def sync_chats(self):
        data = self.get_updates()
        chats = {}
        for update in data["result"]:
            key = update["message"]["chat"]["id"]
            value = update["message"]["from"]["first_name"]
            chats[key] = value
        self.chats = chats
        return chats
    
    def spam(self, text):
        for chat, name in self.chats.items():
            text_message = f"Hello, {name}! {text}"
            self.send_message(chat, text_message)

    def create_excel_chats(self, file_name=None):
        if not file_name:
            now = int(datetime.now().timestamp())
            file_name = f"chats{now}.xlsx"

        self.file_name = file_name
        new_excel = Workbook()
        page = new_excel.active

        chats = self.sync_chats()
        for key, value in chats.items():
            page.append([key, value])
        
        new_excel.save(file_name)
        return file_name
    
    def get_chats_excel(self):
        excel_file = load_workbook(self.file_name)
        page = excel_file[excel_file.sheetnames[0]]
        chats = {}
        for row in page:
            chat_id = row[0].value
            name = row[1].value
            chats[chat_id] = name
        return chats




# response = requests.get(f"{url}getUpdates")
# data = response.json()
# print(data)
# chat_id = 1020812876
# text = "@-->-"
# requests.get(f"{url}SendMessage?chat_id={chat_id}&text={text}")